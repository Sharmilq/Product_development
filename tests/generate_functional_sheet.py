import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
SRC_FILE = "tests/reports/DentNova_Test_Cases.xlsx"
DEST_FILE = "functional-test-cases.xlsx"

def build_functional_sheet():
    print(f"Reading from {SRC_FILE}...")
    wb_src = openpyxl.load_workbook(SRC_FILE)
    
    # Destination workbook
    wb_dest = openpyxl.Workbook()
    ws_dest = wb_dest.active
    ws_dest.title = "Functional Test Cases"
    
    # Target columns
    cols = [
        "TC ID", "Module", "Suite", "Test Case Title", "Preconditions",
        "Test Steps", "Test Data", "Expected Result", "Actual Result",
        "Status", "Priority", "Severity", "Automation Status"
    ]
    
    # Styles
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Navy Blue
    
    font_cell = Font(name="Segoe UI", size=10)
    fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  # Slate 50
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    fill_pass = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # Green 100
    font_pass = Font(name="Segoe UI", size=10, bold=True, color="166534")
    
    fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Red 100
    font_fail = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
    
    fill_skipped = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # Yellow 100
    font_skipped = Font(name="Segoe UI", size=10, bold=True, color="854D0E")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Write headers
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws_dest.cell(row=1, column=c_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws_dest.row_dimensions[1].height = 28
    
    row_count = 2
    
    # Map sheets and read test cases
    # Sheets in SRC_FILE: 'Selenium', 'Appium', 'API', 'Security', 'Load', 'Unit', 'Integration'
    for sheet_name in wb_src.sheetnames:
        if sheet_name == "Traceability":
            continue
        
        ws_src = wb_src[sheet_name]
        print(f"Processing sheet {sheet_name} with {ws_src.max_row - 1} rows...")
        
        # Determine suite name
        if "Selenium" in sheet_name:
            suite = "Web Functional Suite"
        elif "Appium" in sheet_name:
            suite = "Mobile Functional Suite"
        elif "API" in sheet_name:
            suite = "API Integration Suite"
        elif "Security" in sheet_name:
            suite = "Security Controls Suite"
        elif "Load" in sheet_name:
            suite = "Performance Load Suite"
        elif "Unit" in sheet_name:
            suite = "Unit Validation Suite"
        else:
            suite = "System Integration Suite"
            
        for r_idx in range(2, ws_src.max_row + 1):
            # Source columns mapping
            src_id = ws_src.cell(row=r_idx, column=1).value
            src_mod = ws_src.cell(row=r_idx, column=2).value
            src_feat = ws_src.cell(row=r_idx, column=3).value
            src_scenario = ws_src.cell(row=r_idx, column=5).value
            src_pre = ws_src.cell(row=r_idx, column=6).value
            src_steps = ws_src.cell(row=r_idx, column=7).value
            src_exp = ws_src.cell(row=r_idx, column=8).value
            src_act = ws_src.cell(row=r_idx, column=9).value
            src_status = ws_src.cell(row=r_idx, column=10).value
            src_prio = ws_src.cell(row=r_idx, column=11).value
            src_sev = ws_src.cell(row=r_idx, column=12).value
            src_tool = ws_src.cell(row=r_idx, column=13).value
            
            # Map TC ID
            tc_id = src_id
            
            # Map Module (Module + Feature)
            module = f"{src_mod} - {src_feat}" if src_feat else src_mod
            
            # Determine Test Data
            test_data = "N/A"
            if "Login" in module or "Auth" in module:
                test_data = "email: test@dentnova.com\npassword: Test@1234"
            elif "Register" in module:
                test_data = "email: newuser@dentnova.com\npassword: NewPassword123\nname: John Doe"
            elif "OTP" in module:
                test_data = "OTP: 123456\nemail: test@dentnova.com"
            elif "Scan" in module:
                test_data = "image: calculus_tooth.jpg"
            elif "Visit" in module:
                test_data = "date: 26 Jul 2026\ntime: 09:00 AM\nclinic: Clove Dental"
                
            # Determine Automation Status
            auto_status = "Automated" if src_tool and src_tool != "Manual" else "Manual"
            
            # Write mapped row
            row_data = [
                tc_id, module, suite, src_scenario, src_pre,
                src_steps, test_data, src_exp, src_act,
                src_status, src_prio, src_sev, auto_status
            ]
            
            fill_current = fill_even if row_count % 2 == 0 else fill_odd
            
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_dest.cell(row=row_count, column=c_idx, value=val)
                cell.font = font_cell
                cell.fill = fill_current
                cell.border = thin_border
                
                # Alignments
                if c_idx in [1, 10, 11, 12, 13]: # Center columns
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    
                # Status formatting
                if c_idx == 10:
                    if val == "PASS":
                        cell.fill = fill_pass
                        cell.font = font_pass
                    elif val == "FAIL":
                        cell.fill = fill_fail
                        cell.font = font_fail
                    else:
                        cell.fill = fill_skipped
                        cell.font = font_skipped
                        
            ws_dest.row_dimensions[row_count].height = 45
            row_count += 1
            
    # Set column widths
    widths = [14, 25, 25, 35, 30, 45, 25, 40, 40, 12, 10, 10, 16]
    for idx, w in enumerate(widths, 1):
        ws_dest.column_dimensions[get_column_letter(idx)].width = w
        
    ws_dest.freeze_panes = "A2"
    
    print(f"Saving to {DEST_FILE}...")
    wb_dest.save(DEST_FILE)
    print("Functional test cases spreadsheet generated successfully!")

if __name__ == "__main__":
    build_functional_sheet()
