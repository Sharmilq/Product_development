"""
DentNova QA Report Post-Processor
Reads tests/reports/DentNova_Test_Cases.xlsx, computes execution metrics across all tabs,
creates a styled 'Summary' sheet at index 0, and injects visual charts.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

FILE_PATH = "tests/reports/DentNova_Test_Cases.xlsx"

def main():
    print("[*] Reading generated workbook for post-processing...")
    wb = openpyxl.load_workbook(FILE_PATH)

    # 1. Gather stats from each worksheet (skip existing Summary or non-test-case sheets)
    passed_total = 0
    failed_total = 0
    nexec_total = 0
    total_total = 0

    sheet_metrics = {}

    for name in wb.sheetnames:
        if name in ["Traceability Matrix", "Summary"]:
            continue
        
        ws = wb[name]
        passed = 0
        failed = 0
        nexec = 0
        total = 0

        # Scan rows starting from row 2 (row 1 is header)
        for r_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=r_idx, column=10).value # Column J is Status
            if not val:
                continue
            total += 1
            if val == "PASS":
                passed += 1
            elif val == "FAIL":
                failed += 1
            else:
                nexec += 1
        
        sheet_metrics[name] = {
            "Passed": passed,
            "Failed": failed,
            "Not Executed": nexec,
            "Total": total
        }

        passed_total += passed
        failed_total += failed
        nexec_total += nexec
        total_total += total

    print(f"  Passed: {passed_total}")
    print(f"  Failed: {failed_total}")
    print(f"  Not Executed: {nexec_total}")
    print(f"  Total: {total_total}")

    # Calculate Pass Percentage
    # Pass rate of active executed tests = Passed / (Passed + Failed) * 100
    active_executed = passed_total + failed_total
    active_pass_rate = (passed_total / active_executed * 100) if active_executed > 0 else 0.0
    overall_pass_rate = (passed_total / total_total * 100) if total_total > 0 else 0.0

    # 2. Create the Summary Worksheet
    if "Summary" in wb.sheetnames:
        wb.remove(wb["Summary"])
    
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.views.sheetView[0].showGridLines = True

    # Styling elements
    header_fill = PatternFill("solid", fgColor="0F172A") # Slate-900
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True, color="0284C7") # Sky-600
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title block
    ws_sum.cell(row=2, column=2, value="DentNova QA Execution Executive Dashboard").font = title_font
    ws_sum.row_dimensions[2].height = 25

    # Overall Summary Table
    ws_sum.cell(row=4, column=2, value="Execution Metric").font = white_font
    ws_sum.cell(row=4, column=2).fill = header_fill
    ws_sum.cell(row=4, column=2).alignment = left_align
    ws_sum.cell(row=4, column=2).border = thin_border
    
    ws_sum.cell(row=4, column=3, value="Count / Value").font = white_font
    ws_sum.cell(row=4, column=3).fill = header_fill
    ws_sum.cell(row=4, column=3).alignment = center_align
    ws_sum.cell(row=4, column=3).border = thin_border

    metrics = [
        ("Total Tests Configured", total_total),
        ("Passed", passed_total),
        ("Failed", failed_total),
        ("Not Executed", nexec_total),
        ("Active Pass Rate (Executed)", f"{active_pass_rate:.2f}%"),
        ("Overall Progress Coverage", f"{(total_total - nexec_total)/total_total * 100:.2f}%")
    ]

    for offset, (m_name, m_val) in enumerate(metrics, 5):
        cell_a = ws_sum.cell(row=offset, column=2, value=m_name)
        cell_a.font = bold_font if "Rate" in m_name or "Total" in m_name else regular_font
        cell_a.border = thin_border
        cell_a.alignment = left_align

        cell_b = ws_sum.cell(row=offset, column=3, value=m_val)
        cell_b.font = bold_font if "Rate" in m_name or "Total" in m_name else regular_font
        cell_b.border = thin_border
        cell_b.alignment = center_align
        
        # Color codes
        if m_name == "Passed" and m_val > 0:
            cell_b.fill = PatternFill("solid", fgColor="D1FAE5") # light green
        elif m_name == "Failed" and m_val > 0:
            cell_b.fill = PatternFill("solid", fgColor="FEE2E2") # light red
        elif m_name == "Not Executed" and m_val > 0:
            cell_b.fill = PatternFill("solid", fgColor="FEF9C3") # light yellow

    # Module Breakdown Table
    ws_sum.cell(row=13, column=2, value="Module Breakdown").font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    
    headers = ["Test Suite / Worksheet", "Passed", "Failed", "Not Executed", "Total Cases"]
    for col_idx, h in enumerate(headers, 2):
        cell = ws_sum.cell(row=15, column=col_idx, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws_sum.row_dimensions[15].height = 24

    row_offset = 16
    for s_name, counts in sheet_metrics.items():
        cell_name = ws_sum.cell(row=row_offset, column=2, value=s_name)
        cell_name.font = regular_font
        cell_name.border = thin_border
        cell_name.alignment = left_align

        for c_idx, key in enumerate(["Passed", "Failed", "Not Executed", "Total"], 3):
            cell_val = ws_sum.cell(row=row_offset, column=c_idx, value=counts[key])
            cell_val.font = regular_font
            cell_val.border = thin_border
            cell_val.alignment = center_align
            
            # Sub-coloring status columns
            if key == "Passed" and counts[key] > 0:
                cell_val.fill = PatternFill("solid", fgColor="E8F5E9")
            elif key == "Failed" and counts[key] > 0:
                cell_val.fill = PatternFill("solid", fgColor="FFEBEE")
            elif key == "Not Executed" and counts[key] > 0:
                cell_val.fill = PatternFill("solid", fgColor="FFFDE7")
        row_offset += 1

    # Add Breakdown Total Row
    total_row = row_offset
    cell_tot = ws_sum.cell(row=total_row, column=2, value="Total")
    cell_tot.font = bold_font
    cell_tot.border = thin_border
    cell_tot.alignment = left_align
    cell_tot.fill = PatternFill("solid", fgColor="F1F5F9")

    ws_sum.cell(row=total_row, column=3, value=passed_total).font = bold_font
    ws_sum.cell(row=total_row, column=3).border = thin_border
    ws_sum.cell(row=total_row, column=3).fill = PatternFill("solid", fgColor="D1FAE5")
    ws_sum.cell(row=total_row, column=3).alignment = center_align

    ws_sum.cell(row=total_row, column=4, value=failed_total).font = bold_font
    ws_sum.cell(row=total_row, column=4).border = thin_border
    ws_sum.cell(row=total_row, column=4).fill = PatternFill("solid", fgColor="FEE2E2")
    ws_sum.cell(row=total_row, column=4).alignment = center_align

    ws_sum.cell(row=total_row, column=5, value=nexec_total).font = bold_font
    ws_sum.cell(row=total_row, column=5).border = thin_border
    ws_sum.cell(row=total_row, column=5).fill = PatternFill("solid", fgColor="FEF9C3")
    ws_sum.cell(row=total_row, column=5).alignment = center_align

    ws_sum.cell(row=total_row, column=6, value=total_total).font = bold_font
    ws_sum.cell(row=total_row, column=6).border = thin_border
    ws_sum.cell(row=total_row, column=6).fill = PatternFill("solid", fgColor="E2E8F0")
    ws_sum.cell(row=total_row, column=6).alignment = center_align

    # 3. Create the Charts
    # Chart 1: Pie Chart for Overall execution status
    pie = PieChart()
    labels = Reference(ws_sum, min_col=2, min_row=5, max_row=7)
    data = Reference(ws_sum, min_col=3, min_row=4, max_row=7)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "DentNova E2E Test Execution Status Summary"
    pie.width = 16
    pie.height = 11

    # Add the Pie chart to the sheet
    ws_sum.add_chart(pie, "H2")

    # Chart 2: Bar Chart for Suite Breakdown status
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.grouping = "stacked"
    bar.overlap = 100
    bar.title = "Suite Breakdown (Passed vs Not Executed)"
    # Categories: Suite Names (column B, rows 16 to total_row-1)
    cats = Reference(ws_sum, min_col=2, min_row=16, max_row=total_row-1)
    # Data: columns Passed (3), Failed (4), Not Executed (5)
    data_breakdown = Reference(ws_sum, min_col=3, min_row=15, max_col=5, max_row=total_row-1)
    
    bar.add_data(data_breakdown, titles_from_data=True)
    bar.set_categories(cats)
    bar.width = 20
    bar.height = 12

    ws_sum.add_chart(bar, "H20")

    # Format column dimensions for presentation
    ws_sum.column_dimensions['A'].width = 3
    ws_sum.column_dimensions['B'].width = 32
    ws_sum.column_dimensions['C'].width = 16
    ws_sum.column_dimensions['D'].width = 12
    ws_sum.column_dimensions['E'].width = 16
    ws_sum.column_dimensions['F'].width = 16

    wb.save(FILE_PATH)
    print("[+] Dashboard Summary & Charts added successfully!")

if __name__ == "__main__":
    main()
